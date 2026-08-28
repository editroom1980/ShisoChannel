# -*- coding: utf-8 -*-
"""宍粟市の行政区（自治会）ごとの人口・世帯数

なぜ：
  「うちの地区は何人ですか」「西町の人口は」に答えられなかった。
  地区の名前は市内に160ほどあり、市民は町名ではなく自分の地区名で話す。
  ★聞き取りのヒントにも効く（バス停で同じ効果を確かめている）。

どこから：
  市の『人口と世帯数』ページにある「行政区別住民基本台帳人口・世帯数」Excel（毎月更新）。
"""
import json, pathlib, re, ssl, sys, urllib.request
import certifi

根 = pathlib.Path(__file__).resolve().parent.parent
頁 = ("https://www.city.shiso.lg.jp/soshiki/shiminseikatsu/shimin/"
      "tantojoho/jinkoutokei/20001.html")

def 取る(url, 秒=180):
    c = ssl.create_default_context(cafile=certifi.where())
    r = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return urllib.request.urlopen(r, timeout=秒, context=c).read()

def 主():
    try:
        import openpyxl
    except ImportError:
        print("！ openpyxl が無い"); return 1
    h = 取る(頁, 60).decode("utf-8", "replace")
    m = re.search(r'href="(//[^"]*gyouseikubetsujinko[^"]*\.xlsx?)"', h)
    if not m:
        print("！ 行政区別のExcelが頁に無い。頁の作りが変わった可能性"); return 1
    url = "https:" + m.group(1)
    仮 = 根 / "scripts" / "_gyousei.xlsx"
    仮.write_bytes(取る(url))
    w = openpyxl.load_workbook(str(仮), data_only=True)
    s = w[w.sheetnames[0]]
    仮.unlink()
    基準 = ""
    for r in range(1, 8):
        for c in range(2, 6):
            v = s.cell(r, c).value
            if v is not None and "基準日" in str(s.cell(r, c - 1).value or ""):
                基準 = str(v)[:10]
    出 = []
    始 = None
    for r in range(1, s.max_row + 1):
        if str(s.cell(r, 1).value or "").strip() == "行政区名":
            始 = r + 1; break
    if 始 is None:
        print("！ 見出し『行政区名』が見つからない"); return 1
    for r in range(始, s.max_row + 1):
        名 = s.cell(r, 1).value
        if 名 is None: continue
        名 = str(名).strip().replace("　", "")
        # ★町の小計行（山崎・一宮・波賀・千種）を混ぜない（2026-08-28の検算で発覚）。
        #   混ぜると合計がちょうど2倍（65,442人＝32,721×2）になる
        if not 名 or 名 in ("合計", "総計", "計", "宍粟市",
                            "山崎", "一宮", "波賀", "千種"): continue
        try:
            世 = int(s.cell(r, 2).value or 0)
            男 = int(s.cell(r, 3).value or 0)
            女 = int(s.cell(r, 4).value or 0)
            計 = int(s.cell(r, 5).value or 0)
        except (TypeError, ValueError):
            continue
        if 計 <= 0: continue
        出.append({"区": 名, "世帯": 世, "男": 男, "女": 女, "計": 計})
    if not 出:
        print("！ 行政区が1件も読めなかった"); return 1
    合 = sum(x["計"] for x in 出)
    世合 = sum(x["世帯"] for x in 出)
    先 = 根 / "shiso_gyousei.json"
    先.write_text(json.dumps({
        "作成": "市公式『行政区別住民基本台帳人口・世帯数』より",
        "url": 頁, "基準日": 基準, "区の数": len(出),
        "合計の人口": 合, "合計の世帯": 世合,
        "項目": sorted(出, key=lambda x: -x["計"])}, ensure_ascii=False, indent=1),
        encoding="utf-8")
    print(f"○ 行政区 {len(出)}区 → {先}（{先.stat().st_size//1024}KB）")
    print(f"  基準日 {基準}／合計 {合:,}人・{世合:,}世帯")
    # ★検算：市全体の人口（別の資料）と合うか
    j = 根 / "shiso_jinko.json"
    if j.exists():
        d = json.loads(j.read_text(encoding="utf-8"))
        月 = 基準[:7].replace("/", "-") if 基準 else None
        if 月 and 月 in d.get("月ごと", {}):
            市 = d["月ごと"][月]["宍粟市"]["計"]
            印 = "○" if abs(合 - 市) <= 5 else "！ずれている"
            print(f"  検算: 行政区の合計 {合:,}人 / 人口統計の{月} {市:,}人"
                  f"（差 {合-市:+,}人）{印}")
        else:
            新 = d.get("最新")
            print(f"  参考: 人口統計の最新({新}) {d['まとめ']['最新の人口']:,}人")
    大 = 出[:8] if 出 else []
    print("\n  人口の多い行政区:")
    for x in sorted(出, key=lambda x: -x["計"])[:8]:
        print(f"   {x['区']:14s} {x['計']:5,}人 {x['世帯']:4,}世帯")
    小 = [x for x in 出 if x["計"] <= 10]
    print(f"\n  10人以下の行政区 {len(小)}区: {[x['区'] for x in 小][:8]}")
    return 0

if __name__ == "__main__":
    sys.exit(主())
