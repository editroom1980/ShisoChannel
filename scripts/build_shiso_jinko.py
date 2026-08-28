# -*- coding: utf-8 -*-
"""宍粟市の人口の移り変わりを作る（合併前の平成16年度〜現在）

なぜ作るか：
  「人口はどう変わってきましたか」に、今の人数しか答えられなかった（2026-08-28の実測）。
  市長・議員・自治会長に見せる場で、合併からの推移を数で語れることは大きい。

どこから取るか：
  市の『年度別人口統計・異動状況』のExcel（平成16年度〜現在）。
  月ごと・地区ごとに、人口・世帯数・65歳以上・出生死亡・転入転出が全部入っている。
  ★列の位置は年度で違う（古い年度には「65歳以上の世帯数」の列が無い）。
    位置を決め打ちせず、見出しの文字から毎シート決める
"""
import json, re, ssl, sys, pathlib, urllib.request
import certifi, xlrd

根 = pathlib.Path(__file__).resolve().parent.parent
一覧 = "https://www.city.shiso.lg.jp/soshiki/shiminseikatsu/shimin/tantojoho/jinkoutokei/1503879507780.html"
地区ら = ["宍粟市", "山崎", "一宮", "波賀", "千種"]

def _ssl文脈():
    return ssl.create_default_context(cafile=certifi.where())

def 取る(url, 秒=180):
    r = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return urllib.request.urlopen(r, timeout=秒, context=_ssl文脈()).read()

def excelの在処():
    h = 取る(一覧, 60).decode("utf-8", "replace")
    m = re.search(r'href="(//[^"]+\.xlsx?)"', h, re.I)
    if not m:
        raise SystemExit("！ Excelへのリンクが頁に無い。頁の作りが変わった可能性")
    return "https:" + m.group(1)

def 見出しの列(s):
    """3行目・4行目の見出しから、欲しい値がどの列かを決める。
       ★『左記のうち65歳以上』の下の 男・女・計 と、人口の 男・女・計 は
         同じ字なので、上の段（3行目）の区切りを見てから割り当てる"""
    上 = [str(s.cell_value(2, c)).strip() for c in range(s.ncols)]
    下 = [str(s.cell_value(3, c)).strip() for c in range(s.ncols)]
    今 = ""
    枠 = {}
    for c in range(s.ncols):
        if 上[c]:
            今 = 上[c]
        名 = 下[c]
        if not 名:
            continue
        if "65歳" in 今:
            鍵 = {"男": "老男", "女": "老女", "計": "老計"}.get(名)
            if 名.startswith("世帯数"): 鍵 = "老世帯"
        elif "人口" in 今:
            鍵 = {"男": "男", "女": "女", "計": "計", "世帯数": "世帯"}.get(名)
        elif "自然" in 今:
            鍵 = {"出生": "出生", "死亡": "死亡"}.get(名)
        elif "社会" in 今:
            鍵 = {"転入": "転入", "転出": "転出"}.get(名)
        else:
            鍵 = None
        if 鍵 and 鍵 not in 枠:
            枠[鍵] = c
    return 枠

def 数(s, r, c):
    if c is None: return None
    v = s.cell_value(r, c)
    if isinstance(v, float): return int(v)
    v = str(v).replace(",", "").strip()
    return int(v) if re.fullmatch(r"-?\d+", v) else None

def 年月(t):
    """『R8年\n4月末』『H16年\n4月末』を 2026-04 の形に"""
    m = re.search(r"([HRhr])\s*(\d+)\s*年\s*(\d+)\s*月", str(t).replace("\n", ""))
    if not m: return None
    年 = int(m.group(2))
    西 = (1988 + 年) if m.group(1).upper() == "H" else (2018 + 年)
    return f"{西}-{int(m.group(3)):02d}"

# 市の「人口と世帯数（今年度）」の頁。1歳ごとの人数のExcelがここにある
年齢の頁 = ("https://www.city.shiso.lg.jp/soshiki/shiminseikatsu/shimin/"
            "tantojoho/jinkoutokei/20001.html")

def 年齢ごとを取る():
    """1歳きざみの人数（0歳〜110歳超）。
       ★「小学生は何人いますか」に答えるため。年齢の幅は市が決めた区切りではなく、
         学齢（6〜11歳＝小学生）などの暮らしの区切りでこちらが束ねる"""
    try:
        import openpyxl
    except ImportError:
        print("  ！ openpyxl が無いので年齢ごとは飛ばす"); return None
    try:
        h = 取る(年齢の頁, 60).decode("utf-8", "replace")
        m = re.search(r'href="(//[^"]*1saigotojinko[^"]*\.xlsx?)"', h, re.I)
        if not m:
            print("  ！ 1歳ごとのExcelが頁に無い"); return None
        生 = 取る("https:" + m.group(1))
        仮 = 根 / "scripts" / "_1sai.xlsx"
        仮.write_bytes(生)
        w = openpyxl.load_workbook(str(仮), data_only=True)
        s = w[w.sheetnames[0]]
        # 「令和8年3月31日現在」を探す
        いつ = ""
        for r in range(1, 8):
            for c in range(1, 6):
                v = str(s.cell(r, c).value or "")
                if "現在" in v: いつ = v.strip()
        歳 = {}
        for r in range(1, s.max_row + 1):
            a = s.cell(r, 1).value
            if a is None: continue
            t = str(a).strip()
            g = re.match(r"^(\d+)", t)
            if not g: continue
            計 = s.cell(r, 4).value
            if 計 is None: continue
            歳[int(g.group(1))] = {"男": int(s.cell(r, 2).value or 0),
                                   "女": int(s.cell(r, 3).value or 0), "計": int(計)}
        仮.unlink()
        if not 歳: return None
        def 束(始, 終):
            return sum(v["計"] for k, v in 歳.items() if 始 <= k <= 終)
        まとめ = {
            "いつ": いつ,
            "0歳から5歳（就学前）": 束(0, 5),
            "6歳から11歳（小学生）": 束(6, 11),
            "12歳から14歳（中学生）": 束(12, 14),
            "15歳から17歳（高校生）": 束(15, 17),
            "18歳から64歳": 束(18, 64),
            "65歳以上": 束(65, 200),
            "75歳以上": 束(75, 200),
            "90歳以上": 束(90, 200),
            "100歳以上": 束(100, 200),
        }
        # ★検算：年代を足した数が全体と合うか（検収の三原則②「保存量を検算しろ」）。
        #   束ね方を間違えると、どこかの年齢が二重になったり抜けたりする
        足 = (まとめ["0歳から5歳（就学前）"] + まとめ["6歳から11歳（小学生）"]
              + まとめ["12歳から14歳（中学生）"] + まとめ["15歳から17歳（高校生）"]
              + まとめ["18歳から64歳"] + まとめ["65歳以上"])
        全 = sum(v["計"] for v in 歳.values())
        if 足 != 全:
            print(f"  ！ 年代の足し算が合わない（束ねた {足:,} ≠ 全体 {全:,}）")
        else:
            print(f"  年齢の検算: 束ねた{足:,}人 ＝ 表の全体{全:,}人")
        まとめ["合計"] = 全
        print(f"  年齢ごと {len(歳)}区切り（{いつ}）"
              f"／小学生 {まとめ['6歳から11歳（小学生）']:,}人"
              f"・100歳以上 {まとめ['100歳以上']:,}人")
        return {"いつ": いつ, "歳ごと": {str(k): v for k, v in sorted(歳.items())},
                "まとめ": まとめ}
    except Exception as e:
        print(f"  ！ 年齢ごとが取れない: {e}"); return None


def 主():
    url = excelの在処()
    print("Excel:", url)
    生 = 取る(url)
    仮 = 根 / "scripts" / "_nendo.xls"
    仮.write_bytes(生)
    b = xlrd.open_workbook(str(仮))
    仮.unlink()
    全 = {}
    for 名 in b.sheet_names():
        if "集計中" in 名:          # 途中の控えは本表と重なるので使わない
            continue
        s = b.sheet_by_name(名)
        if s.nrows < 6: continue
        枠 = 見出しの列(s)
        if "計" not in 枠:
            print(f"  ！{名}: 人口の列が読めない（見出しの作りが違う）"); continue
        今月 = None
        取 = 0
        for r in range(4, s.nrows):
            ym = 年月(s.cell_value(r, 0))
            if ym: 今月 = ym
            if 今月 is None: continue
            地 = str(s.cell_value(s.row_len(r) and r, 1)).strip()
            if 地 == "全体": 地 = "宍粟市"      # 平成16年度は「全体」表記
            if 地 not in 地区ら: continue
            計 = 数(s, r, 枠.get("計"))
            if not 計: continue                 # 先の月の空欄
            一 = {"男": 数(s, r, 枠.get("男")), "女": 数(s, r, 枠.get("女")),
                  "計": 計, "世帯": 数(s, r, 枠.get("世帯"))}
            for k in ("老計", "出生", "死亡", "転入", "転出"):
                v = 数(s, r, 枠.get(k))
                if v is not None: 一[k] = v
            # ★平成16年度〜17年2月は65歳以上の欄が空で、0が入っている（市の表のまま）。
            #   0のまま残すと「高齢化率0%」という嘘の数になるので捨てる（2026-08-28）
            if 一.get("老計") == 0: 一.pop("老計")
            全.setdefault(今月, {})[地] = 一
            取 += 1
        print(f"  {名:16s} {取:3d}行  列={sorted(枠)}")
    # 5地区そろっている月だけ使う
    全 = {k: v for k, v in 全.items() if len(v) == len(地区ら)}
    if not 全:
        print("！ 1か月も読めなかった"); return 1
    順 = sorted(全)
    古, 新 = 順[0], 順[-1]
    def 番(k): return int(k[:4]) * 12 + int(k[5:])
    欠 = [f"{x//12}-{x%12 or 12:02d}" for a, b2 in zip(順, 順[1:])
          for x in range(番(a) + 1, 番(b2))]
    老 = {k: v for k, v in 全.items() if v["宍粟市"].get("老計", 0) > 0}
    老新 = max(老) if 老 else None
    老古 = min(老) if 老 else None
    出 = {
        "作成": "市公式『年度別人口統計・異動状況』（住民基本台帳）より",
        "url": 一覧, "最古": 古, "最新": 新,
        "月ごと": {k: 全[k] for k in 順},
        "まとめ": {
            "最新の人口": 全[新]["宍粟市"]["計"], "最新の世帯数": 全[新]["宍粟市"]["世帯"],
            "最古の人口": 全[古]["宍粟市"]["計"],
            "増減": 全[新]["宍粟市"]["計"] - 全[古]["宍粟市"]["計"],
            "年数": round((番(新) - 番(古)) / 12, 1),
            "地区ごとの増減": {
                地: {"最古": 全[古][地]["計"], "最新": 全[新][地]["計"],
                     "増減": 全[新][地]["計"] - 全[古][地]["計"],
                     "率": round((全[新][地]["計"] - 全[古][地]["計"]) / 全[古][地]["計"] * 100, 1)}
                for 地 in 地区ら},
        },
    }
    if 老新:
        出["まとめ"]["高齢化率"] = {
            "月": 老新,
            "率": round(全[老新]["宍粟市"]["老計"] / 全[老新]["宍粟市"]["計"] * 100, 1),
            "65歳以上": 全[老新]["宍粟市"]["老計"],
            "地区ごと": {地: round(全[老新][地]["老計"] / 全[老新][地]["計"] * 100, 1)
                        for 地 in 地区ら},
            "最古の月": 老古,
            "最古の率": round(全[老古]["宍粟市"]["老計"] / 全[老古]["宍粟市"]["計"] * 100, 1),
        }
    # ★年ごとの出生・死亡・転入・転出（2026-08-28）。
    #   月の数だけでは「赤ちゃんが減っている」ことが見えない。
    #   実測：出生は2005年355人 → 2025年120人。20年で3分の1になっている。
    #   ★12か月そろった年だけ入れる（途中の年を並べると激減したように見えて嘘になる）
    年ごと = {}
    for k in 順:
        s2 = 全[k]["宍粟市"]
        if "出生" not in s2: continue
        y = k[:4]
        a = 年ごと.setdefault(y, {"出生": 0, "死亡": 0, "転入": 0, "転出": 0, "月数": 0})
        for t in ("出生", "死亡", "転入", "転出"): a[t] += s2.get(t, 0)
        a["月数"] += 1
    年ごと = {y: {"出生": a["出生"], "死亡": a["死亡"], "転入": a["転入"], "転出": a["転出"],
                 "自然増減": a["出生"] - a["死亡"], "社会増減": a["転入"] - a["転出"]}
              for y, a in 年ごと.items() if a["月数"] == 12}
    出["年ごと"] = 年ごと
    if 年ごと:
        年順 = sorted(年ごと)
        出["まとめ"]["出生"] = {
            "新しい年": 年順[-1], "新しい数": 年ごと[年順[-1]]["出生"],
            "古い年": 年順[0], "古い数": 年ごと[年順[0]]["出生"],
            "死亡": 年ごと[年順[-1]]["死亡"],
            "自然増減": 年ごと[年順[-1]]["自然増減"],
        }
    年齢 = 年齢ごとを取る()
    if 年齢: 出["年齢ごと"] = 年齢
    先 = 根 / "shiso_jinko.json"
    先.write_text(json.dumps(出, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\n○ {len(順)}か月分（{古}〜{新}）→ {先}（{先.stat().st_size//1024}KB）")
    print(f"検算: {古}〜{新} は連続で {番(新)-番(古)+1}か月 / 取れた {len(順)}か月／飛び {len(欠)}か月")
    if 欠: print("  取れていない月:", " ".join(欠))
    m = 出["まとめ"]
    print(f"\n{古} {m['最古の人口']:,}人 → {新} {m['最新の人口']:,}人"
          f"（{m['年数']}年で {m['増減']:+,}人／{m['増減']/m['最古の人口']*100:+.1f}%）")
    for 地, v in m["地区ごとの増減"].items():
        print(f"  {地:4s} {v['最古']:6,} → {v['最新']:6,}  {v['増減']:+6,}人 ({v['率']:+.1f}%)")
    if "出生" in m:
        b = m["出生"]
        print(f"\n生まれた赤ちゃん {b['古い年']}年 {b['古い数']}人 → {b['新しい年']}年 {b['新しい数']}人"
              f"（亡くなった方 {b['死亡']}人／自然増減 {b['自然増減']:+}人）")
    if "高齢化率" in m:
        k = m["高齢化率"]
        print(f"\n高齢化率 {k['最古の月']} {k['最古の率']}% → {k['月']} {k['率']}%"
              f"（65歳以上 {k['65歳以上']:,}人）")
        print("  ", "  ".join(f"{地} {r}%" for 地, r in k["地区ごと"].items()))
    return 0

if __name__ == "__main__":
    sys.exit(主())
